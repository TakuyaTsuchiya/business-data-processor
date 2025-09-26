"""
ãƒŸãƒ©ã‚¤ãƒ«ï¼ˆãƒ•ã‚§ã‚¤ã‚¹å°ç­’ï¼‰å·®è¾¼ã¿ç”¨ãƒªã‚¹ãƒˆç”»é¢ãƒ¢ã‚¸ãƒ¥ãƒ¼ãƒ«
å¥‘ç´„è€…ã€ä¿è¨¼äººã€é€£çµ¡äººã®éƒµé€ç”¨ãƒªã‚¹ãƒˆã‚’ä½œæˆã™ã‚‹ç”»é¢
"""
import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from components.common_ui import display_filter_conditions, display_processing_logs
from processors.mirail_notification import process_mirail_notification


def render_mirail_notification(target_type: str, client_pattern: str):
    """ãƒŸãƒ©ã‚¤ãƒ«ï¼ˆãƒ•ã‚§ã‚¤ã‚¹å°ç­’ï¼‰å…±é€šãƒ¬ãƒ³ãƒ€ãƒªãƒ³ã‚°å‡¦ç†"""

    # ã‚¿ã‚¤ãƒˆãƒ«è¨­å®š
    target_name_map = {
        'contractor': 'å¥‘ç´„è€…',
        'guarantor': 'ä¿è¨¼äºº',
        'contact': 'é€£çµ¡äºº'
    }
    target_name = target_name_map.get(target_type, target_type)
    pattern_text = "ï¼ˆ1,4,5ï¼‰" if client_pattern == 'included' else "ï¼ˆ1,4,5ä»¥å¤–ï¼‰"

    st.title("ğŸ“ ãƒŸãƒ©ã‚¤ãƒ« å‚¬å‘Šæ›¸ å·®ã—è¾¼ã¿ãƒªã‚¹ãƒˆ")
    st.subheader(f"{target_name}{pattern_text}ã®ãƒªã‚¹ãƒˆã‚’ä½œæˆ")

    # ãƒ•ã‚£ãƒ«ã‚¿æ¡ä»¶è¡¨ç¤º
    with st.expander("ğŸ“‹ ãƒ•ã‚£ãƒ«ã‚¿æ¡ä»¶", expanded=True):
        base_conditions = [
            "å§”è¨—å…ˆæ³•äººID = 5ã¨ç©ºç™½ã®ã¿é¸æŠ",
            "å…¥é‡‘äºˆå®šæ—¥ < æœ¬æ—¥ï¼ˆæœ¬æ—¥ã‚’é™¤ãï¼‰",
            "å…¥é‡‘äºˆå®šé‡‘é¡ = 2, 3, 5, 12ã‚’é™¤å¤–",
            "å›åãƒ©ãƒ³ã‚¯ â‰  å¼è­·å£«ä»‹å…¥ï¼ˆé™¤å¤–ï¼‰"
        ]

        # ã‚¯ãƒ©ã‚¤ã‚¢ãƒ³ãƒˆCDãƒ•ã‚£ãƒ«ã‚¿
        if client_pattern == 'included':
            base_conditions.append("ã‚¯ãƒ©ã‚¤ã‚¢ãƒ³ãƒˆCD = 1, 4, 5ã‚’é¸æŠ")
        else:
            base_conditions.append("ã‚¯ãƒ©ã‚¤ã‚¢ãƒ³ãƒˆCD â‰  1, 4, 5, 10, 40ï¼ˆã™ã¹ã¦é™¤å¤–ï¼‰")

        # ä½æ‰€ãƒ•ã‚£ãƒ«ã‚¿
        if target_type == 'contractor':
            base_conditions.append("å¥‘ç´„è€…ä½æ‰€ï¼ˆéƒµä¾¿ç•ªå·ãƒ»ç¾ä½æ‰€1ãƒ»2ãƒ»3ï¼‰ãŒã™ã¹ã¦å…¥åŠ›ã•ã‚Œã¦ã„ã‚‹")
        elif target_type == 'guarantor':
            base_conditions.append("ä¿è¨¼äºº1ã¾ãŸã¯ä¿è¨¼äºº2ã®ä½æ‰€ãŒå®Œå…¨ï¼ˆå°‘ãªãã¨ã‚‚ä¸€æ–¹ã®éƒµé€ãŒå¯èƒ½ï¼‰")
        else:  # contact
            base_conditions.append("ç·Šæ€¥é€£çµ¡äºº1ã¾ãŸã¯ç·Šæ€¥é€£çµ¡äºº2ã®ä½æ‰€ãŒå®Œå…¨ï¼ˆå°‘ãªãã¨ã‚‚ä¸€æ–¹ã®éƒµé€ãŒå¯èƒ½ï¼‰")

        display_filter_conditions(base_conditions)

    # ãƒ•ã‚¡ã‚¤ãƒ«ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰
    st.markdown("### ğŸ“‚ ãƒ•ã‚¡ã‚¤ãƒ«ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰")
    uploaded_file = st.file_uploader(
        "ContractList*.csvãƒ•ã‚¡ã‚¤ãƒ«ã‚’é¸æŠã—ã¦ãã ã•ã„",
        type=['csv'],
        key=f"mirail_{target_type}_{client_pattern}"
    )

    if uploaded_file is not None:
        # å‡¦ç†å®Ÿè¡Œãƒœã‚¿ãƒ³
        if st.button("ğŸš€ å‡¦ç†ã‚’å®Ÿè¡Œ", type="primary", key=f"process_mirail_{target_type}_{client_pattern}"):
            with st.spinner("å‡¦ç†ä¸­..."):
                try:
                    # ãƒ•ã‚¡ã‚¤ãƒ«èª­ã¿è¾¼ã¿
                    file_content = uploaded_file.read()

                    # ãƒ—ãƒ­ã‚»ãƒƒã‚µãƒ¼å®Ÿè¡Œ
                    result_df, filename, message, logs = process_mirail_notification(
                        file_content, target_type, client_pattern
                    )

                    # ãƒ­ã‚°è¡¨ç¤º
                    display_processing_logs(logs)

                    # çµæœå‡¦ç†
                    if result_df is not None and len(result_df) > 0:
                        # æˆåŠŸãƒ¡ãƒƒã‚»ãƒ¼ã‚¸
                        st.success(message)

                        # ãƒ‡ãƒ¼ã‚¿ãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼
                        with st.expander(f"ğŸ“Š ãƒ‡ãƒ¼ã‚¿ãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼ï¼ˆå…ˆé ­10ä»¶ï¼‰", expanded=True):
                            st.dataframe(result_df.head(10))

                        # Excelå½¢å¼ã§ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ï¼ˆæ¸¸ã‚´ã‚·ãƒƒã‚¯ 12ptï¼‰
                        output = io.BytesIO()

                        # openpyxlã§ç›´æ¥ãƒ¯ãƒ¼ã‚¯ãƒ–ãƒƒã‚¯ã‚’ä½œæˆ
                        wb = Workbook()
                        ws = wb.active
                        ws.title = 'Sheet1'

                        # ãƒ•ã‚©ãƒ³ãƒˆè¨­å®šï¼ˆæ¸¸ã‚´ã‚·ãƒƒã‚¯ Regular 12ptã€ç½«ç·šãªã—ï¼‰
                        custom_font = Font(
                            name='æ¸¸ã‚´ã‚·ãƒƒã‚¯',  # Yu Gothic / æ¸¸ã‚´ã‚·ãƒƒã‚¯ä½“ / YuGothic
                            size=12,
                            bold=False
                        )

                        # ãƒ˜ãƒƒãƒ€ãƒ¼ã‚’æ›¸ãè¾¼ã¿ï¼ˆãƒ•ã‚©ãƒ³ãƒˆé©ç”¨ï¼‰
                        for col_num, column_title in enumerate(result_df.columns, 1):
                            cell = ws.cell(row=1, column=col_num, value=column_title)
                            cell.font = custom_font

                        # ãƒ‡ãƒ¼ã‚¿ã‚’æ›¸ãè¾¼ã¿ï¼ˆãƒ•ã‚©ãƒ³ãƒˆé©ç”¨ï¼‰
                        for row_num, row_data in enumerate(result_df.values, 2):
                            for col_num, cell_value in enumerate(row_data, 1):
                                # NaNã‚„Noneã®å‡¦ç†
                                if pd.isna(cell_value):
                                    cell_value = ''
                                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                                cell.font = custom_font

                        # ä¿å­˜
                        wb.save(output)
                        output.seek(0)

                        st.download_button(
                            label=f"ğŸ“¥ {filename}ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
                            data=output.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            key=f"download_mirail_{target_type}_{client_pattern}"
                        )
                    else:
                        # ã‚¨ãƒ©ãƒ¼ãƒ¡ãƒƒã‚»ãƒ¼ã‚¸ï¼ˆ0ä»¶ã®å ´åˆï¼‰
                        st.error(message)

                except Exception as e:
                    st.error(f"å‡¦ç†ä¸­ã«ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {str(e)}")


# app.pyã‹ã‚‰å‘¼ã°ã‚Œã‚‹6ã¤ã®é–¢æ•°
def show_mirail_contractor_145():
    """å¥‘ç´„è€…ï¼ˆ1,4,5ï¼‰ç”»é¢"""
    render_mirail_notification('contractor', 'included')


def show_mirail_contractor_not145():
    """å¥‘ç´„è€…ï¼ˆ1,4,5ä»¥å¤–ï¼‰ç”»é¢"""
    render_mirail_notification('contractor', 'excluded')


def show_mirail_guarantor_145():
    """ä¿è¨¼äººï¼ˆ1,4,5ï¼‰ç”»é¢"""
    render_mirail_notification('guarantor', 'included')


def show_mirail_guarantor_not145():
    """ä¿è¨¼äººï¼ˆ1,4,5ä»¥å¤–ï¼‰ç”»é¢"""
    render_mirail_notification('guarantor', 'excluded')


def show_mirail_contact_145():
    """é€£çµ¡äººï¼ˆ1,4,5ï¼‰ç”»é¢"""
    render_mirail_notification('contact', 'included')


def show_mirail_contact_not145():
    """é€£çµ¡äººï¼ˆ1,4,5ä»¥å¤–ï¼‰ç”»é¢"""
    render_mirail_notification('contact', 'excluded')