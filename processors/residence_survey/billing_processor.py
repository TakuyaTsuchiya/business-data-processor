"""
居住訪問調査報告書 請求書作成プロセッサー
Kintoneエクスポートデータから弁護士法人ごとの請求書作成用データを生成
"""
import pandas as pd
import io
from datetime import datetime
from typing import Tuple, List
from openpyxl.styles import Font, PatternFill

# エリア外都道府県リスト
OUT_OF_AREA_PREFECTURES = [
    '鹿児島県', '大分県', '宮崎県', '沖縄県', '石川県',
    '新潟県', '秋田県', '北海道', '青森県', '岩手県', '山形県'
]

# 高橋裕次郎法律事務所の識別名
TAKAHASHI_LAW_FIRM = '弁護士法人高橋裕次郎法律事務所'


def is_out_of_area(address: str) -> bool:
    """住所からエリア外判定を行う"""
    if pd.isna(address):
        return False

    address_str = str(address)
    return any(address_str.startswith(pref) for pref in OUT_OF_AREA_PREFECTURES)


def get_expense_notes(address: str, times: int) -> str:
    """
    費用備考を生成

    Args:
        address: 住所
        times: 調査回数 (1, 2, 3)

    Returns:
        費用備考文字列
    """
    base_text = f"現地調査({times}回目)"

    if is_out_of_area(address):
        return f"{base_text}　エリア外"
    else:
        return base_text


def get_survey_month(survey_date) -> str:
    """
    調査日時から月を判定 (YYYYMM形式)

    Args:
        survey_date: 調査日時（文字列または日付）

    Returns:
        YYYYMM形式の文字列、調査日がない場合はNone
    """
    if pd.isna(survey_date):
        return None

    try:
        date = pd.to_datetime(survey_date)
        return date.strftime('%Y%m')
    except:
        return None


def get_target_month(row: pd.Series) -> str:
    """
    提出日から対象月を判定 (YYYYMM形式)

    Args:
        row: データ行

    Returns:
        YYYYMM形式の文字列、提出日がない場合はNone
    """
    # 3回目 → 2回目 → 1回目の優先順で確認
    for col in ['3回目提出日', '2回目提出日', '1回目提出日']:
        if pd.notna(row[col]):
            try:
                date = pd.to_datetime(row[col])
                return date.strftime('%Y%m')
            except:
                continue

    return None


def determine_billing_rows(row: pd.Series, is_takahashi: bool, selected_month: str = None) -> List[int]:
    """
    請求対象の回数リストを決定

    Args:
        row: データ行
        is_takahashi: 高橋裕次郎法律事務所かどうか
        selected_month: 選択された調査月（YYYYMM形式）。Noneの場合は従来の提出日ベース

    Returns:
        請求対象の回数リスト [1], [1,2], [3], [1,2,3] など
    """
    # 調査月フィルターが指定されている場合
    if selected_month:
        # 各回の提出月を取得（提出日ベースで集計）
        submission_months = {
            1: get_survey_month(row.get('1回目提出日')),
            2: get_survey_month(row.get('2回目提出日')),
            3: get_survey_month(row.get('3回目提出日'))
        }

        # 選択月に該当する提出回のみを抽出
        matching_times = [times for times, month in submission_months.items() if month == selected_month]

        # 該当する調査がない場合
        if not matching_times:
            return []

        # 高橋裕次郎の特例（該当する調査回の中で適用）
        if is_takahashi:
            # 3回目が含まれている場合、該当するすべての回を返す
            if 3 in matching_times:
                return matching_times
            # 2回目が含まれている場合
            elif 2 in matching_times:
                return matching_times
            # 1回目のみ
            elif 1 in matching_times:
                return matching_times
        # 通常パターン
        else:
            # 3回目が含まれている場合、3回目のみ
            if 3 in matching_times:
                return [3]
            # 2回目が含まれている場合、1回目+2回目を請求
            elif 2 in matching_times:
                return [1, 2]
            # 1回目のみ
            elif 1 in matching_times:
                return [1]

        return matching_times

    # 従来の提出日ベースの処理
    else:
        has_1st = pd.notna(row['1回目提出日'])
        has_2nd = pd.notna(row['2回目提出日'])
        has_3rd = pd.notna(row['3回目提出日'])

        # 通常パターン
        if not is_takahashi:
            if has_3rd:
                return [3]  # 3回目のみ
            elif has_2nd:
                return [1, 2]  # 1回目+2回目
            elif has_1st:
                return [1]  # 1回目のみ

        # 高橋裕次郎の特例
        else:
            if has_3rd:
                return [1, 2, 3]  # 全て
            elif has_2nd:
                return [1, 2]
            elif has_1st:
                return [1]

        return []  # 提出日がない場合は請求対象外


def create_billing_row(row: pd.Series, times: int) -> dict:
    """
    請求データ行を生成

    Args:
        row: 元データ行
        times: 調査回数 (1, 2, 3)

    Returns:
        請求データ行の辞書
    """
    return {
        '受託日': '',
        '依頼者債権番号': row['会員番号'] if pd.notna(row['会員番号']) else '',
        '債務者氏名': row['居住者名'] if pd.notna(row['居住者名']) else '',
        '事件種類名称': '',
        '事件名': '',
        '費用発生日': '',
        '費用仕分コード名称': '調査費用',
        '費用コード名称': '',
        '金額': '',
        '費用備考': get_expense_notes(row['住所'], times),
        '業者': 'ミライル'
    }


def process_residence_survey_billing(df: pd.DataFrame, selected_month: str = None) -> Tuple[io.BytesIO, str, str, List[str]]:
    """
    居住訪問調査報告書から請求書作成用データを生成

    Args:
        df: 入力CSV DataFrame
        selected_month: 選択された調査月（YYYYMM形式）。Noneの場合は従来の提出日ベース

    Returns:
        (excel_buffer, filename, message, logs)
    """
    logs = []

    # 入力データ検証
    required_columns = [
        'レコード番号', '依頼元', '会員番号', '居住者名', '住所',
        '調査日時【１回目】', '調査日時【２回目】', '調査日時【３回目】',
        '1回目提出日', '2回目提出日', '3回目提出日', '請求事項'
    ]

    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        raise ValueError(f"必須列が不足しています: {', '.join(missing_cols)}")

    logs.append(f"入力データ: {len(df)}件")
    if selected_month:
        logs.append(f"選択された提出月: {selected_month[:4]}年{selected_month[4:]}月")

    # 弁護士法人ごとにグループ化して処理
    law_firm_data = {}
    target_month = selected_month  # 選択月を優先使用
    skipped_count = 0
    processed_count = 0

    for idx, row in df.iterrows():
        # 調査月フィルター使用時は、従来の提出日チェックをスキップ
        if not selected_month:
            # 提出日がない行はスキップ
            month = get_target_month(row)
            if not month:
                skipped_count += 1
                continue

            # 対象月を記録（最初に見つかった月を使用）
            if target_month is None:
                target_month = month

        # 依頼元の取得
        law_firm = row['依頼元'] if pd.notna(row['依頼元']) else '依頼元不明'

        # 高橋裕次郎かどうか判定
        is_takahashi = law_firm == TAKAHASHI_LAW_FIRM

        # 請求対象の回数を決定
        billing_times = determine_billing_rows(row, is_takahashi, selected_month)

        if not billing_times:
            skipped_count += 1
            continue

        # 弁護士法人ごとのデータを初期化
        if law_firm not in law_firm_data:
            law_firm_data[law_firm] = []

        # 各回数分の請求行を生成
        for times in billing_times:
            billing_row = create_billing_row(row, times)
            law_firm_data[law_firm].append(billing_row)
            processed_count += 1

    logs.append(f"処理済み: {processed_count}行")
    logs.append(f"スキップ: {skipped_count}件")
    logs.append(f"弁護士法人数: {len(law_firm_data)}")

    # 対象月が見つからない場合
    if target_month is None:
        raise ValueError("対象データが見つかりませんでした")

    # Excelファイル作成
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        for law_firm, billing_rows in law_firm_data.items():
            # DataFrameに変換
            df_billing = pd.DataFrame(billing_rows)

            # シート名を作成（Excelの31文字制限を考慮）
            sheet_name = law_firm[:31] if len(law_firm) > 31 else law_firm

            # ヘッダー行を作成
            header_row = pd.DataFrame([{
                '受託日': '受託日',
                '依頼者債権番号': '依頼者債権番号',
                '債務者氏名': '債務者氏名',
                '事件種類名称': '事件種類名称',
                '事件名': '事件名',
                '費用発生日': '費用発生日',
                '費用仕分コード名称': '費用仕分コード名称',
                '費用コード名称': '費用コード名称',
                '金額': '',
                '費用備考': '費用備考',
                '業者': '業者'
            }])

            # 合計行を作成（金額列のみ0、他は空白）
            total_row = pd.DataFrame([{
                '受託日': '',
                '依頼者債権番号': '',
                '債務者氏名': '',
                '事件種類名称': '',
                '事件名': '',
                '費用発生日': '',
                '費用仕分コード名称': '',
                '費用コード名称': '',
                '金額': 0,
                '費用備考': '',
                '業者': ''
            }])

            # 合計行 + ヘッダー行 + データ行 の順で結合
            df_output = pd.concat([total_row, header_row, df_billing], ignore_index=True)

            # Excelに書き込み（ヘッダーなし）
            df_output.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            logs.append(f"  - {law_firm}: {len(billing_rows)}行")

        # フォント設定と背景色を適用
        workbook = writer.book
        font = Font(name='游ゴシック', size=11)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            # 全セルにフォントを適用
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = font

            # ヘッダー行（2行目、A2〜J2）に黄色背景を適用
            for col_idx in range(1, 11):  # A=1 to J=10
                cell = worksheet.cell(row=2, column=col_idx)
                cell.fill = yellow_fill

    excel_buffer.seek(0)

    # ファイル名生成
    filename = f"【居住訪問調査報告書】{target_month}請求内訳.xlsx"

    # 成功メッセージ
    message = f"✅ 請求書データを作成しました（{len(law_firm_data)}法人、{processed_count}行）"

    return excel_buffer, filename, message, logs
