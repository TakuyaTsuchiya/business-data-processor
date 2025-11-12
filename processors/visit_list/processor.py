"""
訪問リスト作成プロセッサー
Business Data Processor

ContractList.csvから訪問スタッフ用の訪問リストExcelを生成
"""

import pandas as pd
import io
from datetime import datetime
from typing import Tuple, List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Border
from processors.common.prefecture_order import get_prefecture_order, extract_prefecture_from_address


class ContractListColumns:
    """ContractList.csv の列番号定義（121列）"""
    MANAGEMENT_NUMBER = 0
    CONTRACT_TYPE = 2
    ENTRUSTED_STATUS = 13
    OCCUPANCY_STATUS = 14
    ARREARS_STATUS = 15
    EVICTION_ACTUAL_COST = 17
    SALES_REP = 19
    CONTRACTOR_NAME = 20
    CONTRACTOR_ADDRESS1 = 23
    CONTRACTOR_ADDRESS2 = 24
    CONTRACTOR_ADDRESS3 = 25
    GUARANTOR1_NAME = 41
    GUARANTOR1_ADDRESS1 = 43
    GUARANTOR1_ADDRESS2 = 44
    GUARANTOR1_ADDRESS3 = 45
    GUARANTOR2_NAME = 48
    GUARANTOR2_ADDRESS1 = 50
    GUARANTOR2_ADDRESS2 = 51
    GUARANTOR2_ADDRESS3 = 52
    CONTACT1_NAME = 55
    CONTACT1_ADDRESS1 = 58
    CONTACT1_ADDRESS2 = 59
    CONTACT1_ADDRESS3 = 60
    CONTACT2_NAME = 62
    CONTACT2_ADDRESS1 = 64
    CONTACT2_ADDRESS2 = 65
    CONTACT2_ADDRESS3 = 66
    ARREARS_BALANCE = 71
    PAYMENT_DUE_DATE = 72
    PAYMENT_DUE_AMOUNT = 73
    MONTHLY_RENT_TOTAL = 84
    COLLECTION_RANK = 86
    CLIENT_CD = 97
    CLIENT_NAME = 98
    DELEGATED_CORP_ID = 118
    DELEGATED_CORP_NAME = 119
    TERMINATION_DATE = 120




class VisitListConfig:
    """訪問リスト作成の設定・定数"""

    # 5種類の人物別の列定義（ContractListの列番号）
    PERSON_TYPES = {
        "contractor": {
            "name": "契約者",
            "sheet_name": "契約者",
            "output_name_col": "契約者氏名",
            "name_col": ContractListColumns.CONTRACTOR_NAME,
            "address1_col": ContractListColumns.CONTRACTOR_ADDRESS1,
            "address2_col": ContractListColumns.CONTRACTOR_ADDRESS2,
            "address3_col": ContractListColumns.CONTRACTOR_ADDRESS3,
        },
        "guarantor1": {
            "name": "保証人1",
            "sheet_name": "保証人1",
            "output_name_col": "保証人１氏名",
            "name_col": ContractListColumns.GUARANTOR1_NAME,
            "address1_col": ContractListColumns.GUARANTOR1_ADDRESS1,
            "address2_col": ContractListColumns.GUARANTOR1_ADDRESS2,
            "address3_col": ContractListColumns.GUARANTOR1_ADDRESS3,
        },
        "guarantor2": {
            "name": "保証人2",
            "sheet_name": "保証人2",
            "output_name_col": "保証人２氏名",
            "name_col": ContractListColumns.GUARANTOR2_NAME,
            "address1_col": ContractListColumns.GUARANTOR2_ADDRESS1,
            "address2_col": ContractListColumns.GUARANTOR2_ADDRESS2,
            "address3_col": ContractListColumns.GUARANTOR2_ADDRESS3,
        },
        "contact1": {
            "name": "連絡人1",
            "sheet_name": "連絡人1",
            "output_name_col": "緊急連絡人１氏名",
            "name_col": ContractListColumns.CONTACT1_NAME,
            "address1_col": ContractListColumns.CONTACT1_ADDRESS1,
            "address2_col": ContractListColumns.CONTACT1_ADDRESS2,
            "address3_col": ContractListColumns.CONTACT1_ADDRESS3,
        },
        "contact2": {
            "name": "連絡人2",
            "sheet_name": "連絡人2",
            "output_name_col": "緊急連絡人２氏名",
            "name_col": ContractListColumns.CONTACT2_NAME,
            "address1_col": ContractListColumns.CONTACT2_ADDRESS1,
            "address2_col": ContractListColumns.CONTACT2_ADDRESS2,
            "address3_col": ContractListColumns.CONTACT2_ADDRESS3,
        },
    }

    OUTPUT_FILE_PREFIX = "訪問リスト"


def combine_address(address1, address2, address3) -> str:
    """
    住所を結合

    Args:
        address1: 現住所1
        address2: 現住所2
        address3: 現住所3

    Returns:
        str: 結合された住所
    """
    parts = []
    for addr in [address1, address2, address3]:
        if pd.notna(addr) and str(addr).strip() != '':
            parts.append(str(addr).strip())
    return ''.join(parts)


def filter_records(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    """
    フィルタリング処理

    7つのフィルタ条件:
    1. 回収ランク: 「交渉困難」「死亡決定」「弁護士介入」を除外
    2. 入金予定日: 当日以前または空白
    3. 入金予定金額: 2, 3, 5を除外
    4. 委託先法人ID: 5と空白のみ
    5. 滞納残債: 1円以上
    6. クライアントCD: 9268を除外
    7. 受託状況: 「契約中」「契約中(口振停止)」

    Args:
        df: ContractList DataFrame

    Returns:
        Tuple[pd.DataFrame, List[str]]: フィルタ後のDataFrameとログ
    """
    logs = []
    initial_count = len(df)
    logs.append(f"入力レコード数: {initial_count}件")

    # 1. 回収ランクフィルタ（「交渉困難」「死亡決定」「弁護士介入」を除外）
    before_count = len(df)
    mask_rank = ~df.iloc[:, ContractListColumns.COLLECTION_RANK].isin(["交渉困難", "死亡決定", "弁護士介入"])
    df_filtered = df[mask_rank].copy()
    excluded_count = before_count - len(df_filtered)
    logs.append(f"回収ランクフィルタ: {excluded_count}件除外 → {len(df_filtered)}件残存")

    # 2. 入金予定日フィルタ（当日以前または空白）
    before_count = len(df_filtered)
    today = datetime.now().date()
    df_filtered['payment_date_parsed'] = pd.to_datetime(
        df_filtered.iloc[:, ContractListColumns.PAYMENT_DUE_DATE], errors='coerce'
    )
    mask_date = df_filtered['payment_date_parsed'].isna() | (
        df_filtered['payment_date_parsed'].dt.date <= today
    )
    df_filtered = df_filtered[mask_date].copy()
    df_filtered = df_filtered.drop(columns=['payment_date_parsed'])
    excluded_count = before_count - len(df_filtered)
    logs.append(f"入金予定日フィルタ: {excluded_count}件除外 → {len(df_filtered)}件残存")

    # 3. 入金予定金額フィルタ（2, 3, 5を除外）
    before_count = len(df_filtered)
    payment_amount_numeric = pd.to_numeric(df_filtered.iloc[:, ContractListColumns.PAYMENT_DUE_AMOUNT], errors='coerce')
    mask_amount = ~payment_amount_numeric.isin([2, 3, 5])
    df_filtered = df_filtered[mask_amount].copy()
    excluded_count = before_count - len(df_filtered)
    logs.append(f"入金予定金額フィルタ: {excluded_count}件除外 → {len(df_filtered)}件残存")

    # 4. 委託先法人IDフィルタ（5と空白のみ）
    before_count = len(df_filtered)
    delegated_corp_id = df_filtered.iloc[:, ContractListColumns.DELEGATED_CORP_ID]
    mask_corp_id = (delegated_corp_id == '5') | (delegated_corp_id == 5) | delegated_corp_id.isna() | (delegated_corp_id == '')
    df_filtered = df_filtered[mask_corp_id].copy()
    excluded_count = before_count - len(df_filtered)
    logs.append(f"委託先法人IDフィルタ: {excluded_count}件除外 → {len(df_filtered)}件残存")

    # 5. 滞納残債フィルタ（1円以上）
    before_count = len(df_filtered)
    debt_amount = pd.to_numeric(
        df_filtered.iloc[:, ContractListColumns.ARREARS_BALANCE].astype(str).str.replace(',', ''),
        errors='coerce'
    )
    mask_debt = debt_amount >= 1
    df_filtered = df_filtered[mask_debt].copy()
    excluded_count = before_count - len(df_filtered)
    logs.append(f"滞納残債1円以上フィルタ: {excluded_count}件除外 → {len(df_filtered)}件残存")

    # 6. クライアントCDフィルタ（9268を除外）
    before_count = len(df_filtered)
    client_cd = pd.to_numeric(df_filtered.iloc[:, ContractListColumns.CLIENT_CD], errors='coerce')
    mask_client = client_cd != 9268
    df_filtered = df_filtered[mask_client].copy()
    excluded_count = before_count - len(df_filtered)
    logs.append(f"クライアントCD9268除外: {excluded_count}件除外 → {len(df_filtered)}件残存")

    # 7. 受託状況フィルタ（「契約中」「契約中(口振停止)」）
    before_count = len(df_filtered)
    status_values = df_filtered.iloc[:, ContractListColumns.ENTRUSTED_STATUS].astype(str).str.strip()
    mask_status = status_values.isin(["契約中", "契約中(口振停止)"])
    df_filtered = df_filtered[mask_status].copy()
    excluded_count = before_count - len(df_filtered)
    logs.append(f"受託状況フィルタ: {excluded_count}件除外 → {len(df_filtered)}件残存（最終）")

    return df_filtered, logs


def create_output_row(row: pd.Series, person_type: str, config: Dict) -> Dict:
    """
    ContractListの1行から22列の出力データを作成（ラッパー関数）

    Args:
        row: ContractListの1行
        person_type: 人物種類（"contractor", "guarantor1" など）
        config: 人物別の列定義

    Returns:
        Dict: 22列のデータ辞書
    """
    # 単一行をDataFrameに変換してバルク処理を呼び出す
    df_single = pd.DataFrame([row])
    df_output = create_output_row_bulk(df_single, person_type, config)
    return df_output.iloc[0].to_dict()


def create_output_row_bulk(df_person: pd.DataFrame, person_type: str, config: Dict) -> pd.DataFrame:
    """
    DataFrameの全行を一括処理（ベクトル化版）

    Args:
        df_person: ContractListのDataFrame
        person_type: 人物種類（"contractor", "guarantor1" など）
        config: 人物別の列定義

    Returns:
        pd.DataFrame: 22列の出力DataFrame
    """
    # 人物情報の列番号
    name_col = config["name_col"]
    addr1_col = config["address1_col"]
    addr2_col = config["address2_col"]
    addr3_col = config["address3_col"]

    # 結合住所を一括生成
    combined_addresses = (
        df_person.iloc[:, addr1_col].fillna('').astype(str) +
        df_person.iloc[:, addr2_col].fillna('').astype(str) +
        df_person.iloc[:, addr3_col].fillna('').astype(str)
    )

    # 数値列を一括変換
    arrears_balance = pd.to_numeric(
        df_person.iloc[:, ContractListColumns.ARREARS_BALANCE].astype(str).str.replace(',', ''), errors='coerce'
    )
    monthly_rent = pd.to_numeric(
        df_person.iloc[:, ContractListColumns.MONTHLY_RENT_TOTAL].astype(str).str.replace(',', ''), errors='coerce'
    )

    # 滞納月数を計算（滞納残債 ÷ 月額賃料合計、小数点第1位まで）
    # 月額賃料が0またはNaNの場合はNaNを返す
    arrears_months = (arrears_balance / monthly_rent).round(1)

    # 全列を一括取得（ベクトル化）
    df_output = pd.DataFrame({
        "管理番号": df_person.iloc[:, ContractListColumns.MANAGEMENT_NUMBER].values,
        "最新契約種類": df_person.iloc[:, ContractListColumns.CONTRACT_TYPE].values,
        "受託状況": df_person.iloc[:, ContractListColumns.ENTRUSTED_STATUS].values,
        "入居ステータス": df_person.iloc[:, ContractListColumns.OCCUPANCY_STATUS].values,
        "滞納ステータス": df_person.iloc[:, ContractListColumns.ARREARS_STATUS].values,
        "退去手続き（実費）": pd.to_numeric(
            df_person.iloc[:, ContractListColumns.EVICTION_ACTUAL_COST].astype(str).str.replace(',', ''), errors='coerce'
        ),
        "営業担当者": df_person.iloc[:, ContractListColumns.SALES_REP].values,
        config["output_name_col"]: df_person.iloc[:, name_col].values,
        "": combined_addresses.values,
        "現住所1": df_person.iloc[:, addr1_col].values,
        "現住所2": df_person.iloc[:, addr2_col].values,
        "現住所3": df_person.iloc[:, addr3_col].values,
        "滞納残債": arrears_balance,
        "入金予定日": df_person.iloc[:, ContractListColumns.PAYMENT_DUE_DATE].values,
        "入金予定金額": pd.to_numeric(
            df_person.iloc[:, ContractListColumns.PAYMENT_DUE_AMOUNT].astype(str).str.replace(',', ''), errors='coerce'
        ),
        "滞納月数": arrears_months,
        "月額賃料合計": monthly_rent,
        "回収ランク": df_person.iloc[:, ContractListColumns.COLLECTION_RANK].values,
        "クライアントCD": df_person.iloc[:, ContractListColumns.CLIENT_CD].values,
        "クライアント名": df_person.iloc[:, ContractListColumns.CLIENT_NAME].values,
        "委託先法人ID": df_person.iloc[:, ContractListColumns.DELEGATED_CORP_ID].values,
        "委託先法人名": df_person.iloc[:, ContractListColumns.DELEGATED_CORP_NAME].values,
        "解約日": df_person.iloc[:, ContractListColumns.TERMINATION_DATE].values,
    })

    return df_output


def sort_by_prefecture(df: pd.DataFrame) -> pd.DataFrame:
    """
    都道府県順（北→南）→現住所2順でソート（ベクトル化版）

    Args:
        df: 22列形式のDataFrame

    Returns:
        pd.DataFrame: ソート後のDataFrame
    """
    from processors.common.prefecture_order import PREFECTURE_ORDER

    # 都道府県名→順序番号のマッピング辞書を作成
    prefecture_map = {pref: idx for idx, pref in enumerate(PREFECTURE_ORDER)}

    # 現住所1から都道府県を一括抽出（ベクトル化）
    addresses = df['現住所1'].astype(str)

    # 正規表現で都道府県名を抽出（最初にマッチしたもの）
    # 北海道|青森県|...のパターンを作成
    prefecture_pattern = '|'.join(PREFECTURE_ORDER)
    extracted_prefectures = addresses.str.extract(f'({prefecture_pattern})', expand=False)

    # 都道府県名を順序番号にマッピング（見つからない場合は999）
    df['prefecture_order'] = extracted_prefectures.map(prefecture_map).fillna(999).astype(int)

    # 都道府県順 → 現住所2順の2段階ソート
    df_sorted = df.sort_values(['prefecture_order', '現住所2']).drop(columns=['prefecture_order'])

    return df_sorted


def generate_excel(
    df_dict: Dict[str, pd.DataFrame],
    output_filename: str
) -> Tuple[bytes, List[str]]:
    """
    5シートのExcelファイルを生成（フォント設定付き）

    Args:
        df_dict: 人物タイプ別のDataFrame辞書
        output_filename: 出力ファイル名

    Returns:
        Tuple[bytes, List[str]]: Excelバイト列とログ
    """
    logs = []

    # Excelファイルを作成
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        for person_type, config in VisitListConfig.PERSON_TYPES.items():
            if person_type in df_dict and len(df_dict[person_type]) > 0:
                df = df_dict[person_type]
                sheet_name = config["sheet_name"]

                # Excelシートに書き込み
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                logs.append(f"{sheet_name}シート: {len(df)}件")

        # フォント設定と罫線削除を適用
        workbook = writer.book
        font = Font(name='游ゴシック Regular', size=11)

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            # フォントと罫線の設定
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.border = Border()  # 罫線削除

            # 数値列にカンマ区切り書式を適用
            headers = [cell.value for cell in ws[1]]
            numeric_cols = []
            decimal_cols = []
            for idx, header in enumerate(headers, start=1):
                if header in ["退去手続き（実費）", "滞納残債", "入金予定金額", "月額賃料合計"]:
                    numeric_cols.append(idx)
                elif header == "滞納月数":
                    decimal_cols.append(idx)

            # 2行目以降の数値列にカンマ書式
            for row in ws.iter_rows(min_row=2):
                for col_idx in numeric_cols:
                    cell = row[col_idx - 1]
                    # NaNチェック: pd.isna()で判定し、NaNの場合は空文字に置き換え
                    if cell.value is not None:
                        if isinstance(cell.value, (int, float)):
                            import math
                            if math.isnan(cell.value):
                                cell.value = ''
                            else:
                                cell.number_format = '#,##0'

                # 小数点列に小数点第1位フォーマット
                for col_idx in decimal_cols:
                    cell = row[col_idx - 1]
                    if cell.value is not None:
                        if isinstance(cell.value, (int, float)):
                            import math
                            if math.isnan(cell.value):
                                cell.value = ''
                            else:
                                cell.number_format = '0.0'

    excel_buffer.seek(0)
    return excel_buffer.getvalue(), logs


def process_visit_list(df_input: pd.DataFrame) -> Tuple[bytes, str, str, List[str]]:
    """
    訪問リスト作成のメイン処理

    Args:
        df_input: ContractList DataFrame

    Returns:
        Tuple[bytes, str, str, List[str]]: Excelバイト列、ファイル名、メッセージ、ログ
    """
    logs = []

    # 1. フィルタリング
    df_filtered, filter_logs = filter_records(df_input)
    logs.extend(filter_logs)

    if len(df_filtered) == 0:
        return None, "", "フィルタ条件に一致するレコードがありませんでした", logs

    # 2. 5つの人物タイプ別にデータ作成
    df_dict = {}
    for person_type, config in VisitListConfig.PERSON_TYPES.items():
        # 該当人物の現住所1が存在するレコードのみ
        address1_col = config["address1_col"]
        mask = df_filtered.iloc[:, address1_col].notna() & (df_filtered.iloc[:, address1_col] != '')
        df_person = df_filtered[mask].copy()

        if len(df_person) > 0:
            # 一括処理（ベクトル化）
            df_output = create_output_row_bulk(df_person, person_type, config)

            # 都道府県順でソート
            df_output = sort_by_prefecture(df_output)

            df_dict[person_type] = df_output

    # 3. Excelファイル生成
    today = datetime.now()
    filename = f"{today.strftime('%m%d')}訪問リスト.xlsx"
    excel_bytes, excel_logs = generate_excel(df_dict, filename)
    logs.extend(excel_logs)

    # 4. サマリーメッセージ
    total_records = sum(len(df) for df in df_dict.values())
    message = f"訪問リスト作成完了 - 合計 {total_records}件"

    return excel_bytes, filename, message, logs
