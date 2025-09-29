"""
共通UIコンポーネント
Business Data Processor

全画面で使用される共通のUI表示関数をまとめたモジュール
"""

import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font


def safe_dataframe_display(df: pd.DataFrame):
    """安全なDataFrame表示関数（空列重複エラー対応）"""
    # DataFrameのコピーを作成して空列問題を回避
    df_display = df.copy()
    
    # 空文字列のカラム名に一時的な名前を付ける
    columns = list(df_display.columns)
    empty_col_counter = 1
    new_columns = []
    for col in columns:
        if col == "":
            new_columns.append(f"空列{empty_col_counter}")
            empty_col_counter += 1
        else:
            new_columns.append(col)
    
    # 一時的なカラム名を設定して表示
    df_display.columns = new_columns
    return st.dataframe(df_display)


def safe_csv_download(df: pd.DataFrame, filename: str, label: str = "📥 CSVファイルをダウンロード"):
    """安全なCSVダウンロード関数（cp932エンコーディングエラー対応）"""
    # DataFrameのコピーを作成して空列問題を回避
    df_copy = df.copy()
    
    # 空文字列のカラム名に一時的な名前を付ける
    columns = list(df_copy.columns)
    empty_col_counter = 1
    for i, col in enumerate(columns):
        if col == "":
            columns[i] = f"_empty_col_{empty_col_counter}_"
            empty_col_counter += 1
    
    # 一時的なカラム名を設定
    df_copy.columns = columns
    
    try:
        # CSVとして出力する際に元のカラム名に戻す
        csv_data = df_copy.to_csv(index=False, encoding='cp932', errors='replace', header=list(df.columns))
        csv_bytes = csv_data.encode('cp932', errors='replace')
    except UnicodeEncodeError:
        # cp932でエラーが出る場合はUTF-8で出力
        csv_data = df_copy.to_csv(index=False, encoding='utf-8-sig', header=list(df.columns))
        csv_bytes = csv_data.encode('utf-8-sig')
        st.warning("⚠️ 一部の文字がcp932に対応していないため、UTF-8で出力します")
    
    return st.download_button(
        label=label,
        data=csv_bytes,
        file_name=filename,
        mime="text/csv",
        type="primary"
    )


def display_processing_logs(logs: list, title: str = "📊 処理ログ", expanded: bool = False):
    """処理ログの統一表示関数"""
    with st.expander(title, expanded=expanded):
        for log in logs:
            # セパレータ行の特別処理
            if log.startswith("="):
                st.markdown(f"**{log}**")
            # サマリー行の特別処理
            elif log.startswith("【") and log.endswith("】"):
                st.markdown(f"**{log}**")
            # インデントされたサブアイテム
            elif log.startswith("- "):
                st.markdown(f"  • {log}")
            # 通常のログ
            else:
                st.markdown(f"• {log}")


def safe_excel_download(df: pd.DataFrame, filename: str, label: str = "📥 Excelファイルをダウンロード"):
    """安全なExcelダウンロード関数（游ゴシック 12ptフォント適用）"""
    output = io.BytesIO()

    # openpyxlで直接ワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # フォント設定（游ゴシック Regular 12pt、罫線なし）
    custom_font = Font(
        name='游ゴシック',
        size=12,
        bold=False
    )

    # ヘッダーを書き込み（フォント適用）
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = custom_font

    # データを書き込み（フォント適用）
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            # NaNやNoneの処理
            if pd.isna(cell_value):
                cell_value = ''
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.font = custom_font

    # 保存
    wb.save(output)
    output.seek(0)

    return st.download_button(
        label=label,
        data=output.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )


def display_filter_conditions(conditions: list, title: str = "**フィルタ条件:**"):
    """フィルタ条件の統一表示関数"""
    st.markdown(title)
    st.markdown('<div class="filter-condition">', unsafe_allow_html=True)
    for condition in conditions:
        st.markdown(f"• {condition}")
    st.markdown('</div>', unsafe_allow_html=True)