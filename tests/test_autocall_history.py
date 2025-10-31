"""
オートコール履歴プロセッサのユニットテスト
Business Data Processor

TDDで実装: Red-Green-Refactorサイクル
"""

import pytest
import pandas as pd
from datetime import datetime


class TestAutocallHistoryProcessor:
    """AutocallHistoryProcessorクラスのテスト"""

    def test_fillna_最終架電日の空白をフォワードフィル(self, sample_autocall_history_data):
        """最終架電日が空白の場合、前行の値でフォワードフィルされることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)

        # 元データで3行目（10003）の最終架電日は空白だったが、処理後はフィルされている
        # 期待: 2行目（10002は通話済で除外されているので、結果の2行目は10003）の日時が2行目の値を継承
        assert result_df.iloc[1]['交渉日時'] == '2025-10-31 10:15:01'

    def test_filter_通話済を除外(self, sample_autocall_history_data):
        """「架電結果」が「通話済」のレコードが除外されることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)

        # 元データ5行中、通話済が2行（管理番号10002, 10005）
        # 期待: 3行のみ残る
        assert len(result_df) == 3
        assert 10002 not in result_df['管理番号'].values
        assert 10005 not in result_df['管理番号'].values

    def test_mapping_10列の出力形式(self, sample_autocall_history_data):
        """出力が10列のNegotiatesInfo形式であることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)

        # 10列であることを確認
        expected_columns = [
            "管理番号", "交渉日時", "担当", "相手", "手段",
            "回収ランク", "結果", "入金予定日", "予定金額", "交渉備考"
        ]
        assert list(result_df.columns) == expected_columns

    def test_mapping_固定値が正しく設定される(self, sample_autocall_history_data):
        """固定値（担当、手段、結果など）が正しく設定されることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="保証人")
        result_df = processor.process(sample_autocall_history_data)

        # 固定値の確認
        assert all(result_df['担当'] == '')
        assert all(result_df['相手'] == '保証人')
        assert all(result_df['手段'] == '架電')
        assert all(result_df['回収ランク'] == '')
        assert all(result_df['結果'] == 'その他')
        assert all(result_df['入金予定日'] == '')
        assert all(result_df['予定金額'] == '')

    def test_mapping_交渉備考の文字列結合(self, sample_autocall_history_data):
        """交渉備考が「架電番号{架電番号}オートコール　残債{残債}円」形式で結合されることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)

        # 1行目の確認
        first_row = result_df.iloc[0]
        expected_remark = "架電番号090-1234-5678オートコール　残債10000円"
        assert first_row['交渉備考'] == expected_remark

        # 2行目の確認（管理番号10003）
        second_row = result_df.iloc[1]
        expected_remark2 = "架電番号090-3333-4444オートコール　残債20000円"
        assert second_row['交渉備考'] == expected_remark2

    def test_complete_processing(self, sample_autocall_history_data, expected_autocall_history_output):
        """全体処理が期待通りに動作することを確認（統合テスト）"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)

        # 期待されるDataFrameと比較
        pd.testing.assert_frame_equal(
            result_df.reset_index(drop=True),
            expected_autocall_history_output.reset_index(drop=True),
            check_dtype=False
        )


class TestOutputGeneration:
    """出力ファイル名生成のテスト"""

    def test_generate_filename_MMDD形式_xlsx(self):
        """出力ファイル名がMMDD形式（xlsx）であることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="契約者")
        filename = processor.generate_output_filename(extension='xlsx')

        # ファイル名が「MMDDオートコール履歴.xlsx」形式であることを確認
        assert filename.endswith("オートコール履歴.xlsx")

        # MMDDが4桁の数字であることを確認
        mmdd_part = filename.split("オートコール履歴")[0]
        assert len(mmdd_part) == 4
        assert mmdd_part.isdigit()

        # 現在の日付のMMDDと一致することを確認
        today = datetime.now()
        expected_mmdd = today.strftime('%m%d')
        assert mmdd_part == expected_mmdd

    def test_generate_filename_MMDD形式_csv(self):
        """出力ファイル名がMMDD形式（csv）であることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor

        processor = AutocallHistoryProcessor(target_person="契約者")
        filename = processor.generate_output_filename(extension='csv')

        # ファイル名が「MMDDオートコール履歴.csv」形式であることを確認
        assert filename.endswith("オートコール履歴.csv")


class TestExcelGeneration:
    """Excel生成のテスト"""

    def test_generate_excel_正常系(self, sample_autocall_history_data):
        """Excelファイルが正常に生成されることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor
        import openpyxl
        import io

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)
        excel_bytes, logs = processor.generate_excel(result_df)

        # バイト列が返されることを確認
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # ログが返されることを確認
        assert isinstance(logs, list)
        assert len(logs) > 0

        # Excelファイルとして読み込めることを確認
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        assert "オートコール履歴" in wb.sheetnames

    def test_generate_excel_列幅設定(self, sample_autocall_history_data):
        """列幅が正しく設定されることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor
        import openpyxl
        import io

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)
        excel_bytes, logs = processor.generate_excel(result_df)

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["オートコール履歴"]

        # 列幅を確認
        assert ws.column_dimensions['A'].width == 12  # 管理番号
        assert ws.column_dimensions['B'].width == 20  # 交渉日時
        assert ws.column_dimensions['J'].width == 50  # 交渉備考

    def test_generate_excel_フォント設定(self, sample_autocall_history_data):
        """フォントが正しく設定されることを確認"""
        from processors.autocall_history import AutocallHistoryProcessor
        import openpyxl
        import io

        processor = AutocallHistoryProcessor(target_person="契約者")
        result_df = processor.process(sample_autocall_history_data)
        excel_bytes, logs = processor.generate_excel(result_df)

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["オートコール履歴"]

        # 最初のセルのフォントを確認
        cell = ws["A1"]
        assert cell.font.name == "游ゴシック Regular"
        assert cell.font.size == 11
