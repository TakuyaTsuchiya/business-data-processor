#!/usr/bin/env python3
"""
列幅調整のテスト
"""
import pandas as pd
import sys
from pathlib import Path
from openpyxl import load_workbook

# プロジェクトルートをパスに追加
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from processors.residence_survey.billing_processor import process_residence_survey_billing

# サンプルデータのパス
CSV_PATH = "/Users/tchytky/Downloads/居住訪問調査報告書_20251008T152416+0900.csv"

def main():
    print("=" * 80)
    print("列幅調整テスト")
    print("=" * 80)
    print()

    # CSVファイル読み込み
    try:
        df = pd.read_csv(CSV_PATH, encoding='cp932')
    except:
        try:
            df = pd.read_csv(CSV_PATH, encoding='shift_jis')
        except:
            df = pd.read_csv(CSV_PATH, encoding='utf-8-sig')

    print(f"✅ データ読み込み完了: {len(df)}件")
    print()

    # 処理実行（2025年10月を選択）
    selected_month = "202510"
    print(f"選択月: {selected_month[:4]}年{selected_month[4:]}月")
    print()

    excel_buffer, filename, message, logs = process_residence_survey_billing(
        df,
        selected_month=selected_month
    )

    print(message)
    print()

    # Excelファイルを保存
    output_path = f"/Users/tchytky/Desktop/{filename}"
    with open(output_path, 'wb') as f:
        f.write(excel_buffer.read())

    print(f"✅ 出力ファイル: {output_path}")
    print()

    # 列幅を確認
    workbook = load_workbook(output_path)
    first_sheet = workbook.sheetnames[0]
    worksheet = workbook[first_sheet]

    print(f"【{first_sheet}シートの列幅】")
    print(f"  A列（受託日）: {worksheet.column_dimensions['A'].width}")
    print(f"  B列（依頼者債権番号）: {worksheet.column_dimensions['B'].width}")
    print(f"  C列（債務者氏名）: {worksheet.column_dimensions['C'].width} ← 広く設定")
    print(f"  D列（事件種類名称）: {worksheet.column_dimensions['D'].width}")
    print(f"  J列（費用備考）: {worksheet.column_dimensions['J'].width} ← 広く設定")
    print(f"  K列（業者）: {worksheet.column_dimensions['K'].width}")
    print(f"  L列（提出日）: {worksheet.column_dimensions['L'].width}")
    print()

    # サンプルデータを確認
    print("【債務者氏名のサンプル】")
    for row_idx in range(3, 8):  # データ行の最初の5行
        cell = worksheet.cell(row=row_idx, column=3)
        print(f"  {row_idx-2}行目: {cell.value}")
    print()

    print("【費用備考のサンプル】")
    for row_idx in range(3, 8):  # データ行の最初の5行
        cell = worksheet.cell(row=row_idx, column=10)
        print(f"  {row_idx-2}行目: {cell.value}")
    print()

    if worksheet.column_dimensions['C'].width == 25 and worksheet.column_dimensions['J'].width == 25:
        print("✅ 列幅が正しく設定されました")
    else:
        print("❌ 列幅の設定に問題があります")

    print()
    print("=" * 80)
    print("テスト完了")
    print("=" * 80)


if __name__ == "__main__":
    main()
