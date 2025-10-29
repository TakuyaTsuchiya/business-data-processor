"""
訪問リスト作成プロセッサのユニットテスト
Business Data Processor

TDDで実装: Red-Green-Refactorサイクル
"""

import pytest
import pandas as pd
from datetime import datetime, timedelta
import io


class TestAddressCombination:
    """住所結合処理のテスト"""

    def test_combine_address_3つ全て結合(self):
        """現住所1+2+3がすべて存在する場合、正しく結合されることを確認"""
        from processors.visit_list.processor import combine_address
        result = combine_address("北海道", "札幌市", "北区1-1-1")
        assert result == "北海道札幌市北区1-1-1"

    def test_combine_address_現住所3が空白(self):
        """現住所3が空白の場合、現住所1+2のみ結合されることを確認"""
        from processors.visit_list.processor import combine_address
        result = combine_address("北海道", "札幌市", "")
        assert result == "北海道札幌市"

    def test_combine_address_すべて空白(self):
        """すべて空白の場合、空文字が返されることを確認"""
        from processors.visit_list.processor import combine_address
        result = combine_address("", "", "")
        assert result == ""


class TestPrefectureOrder:
    """都道府県順序のテスト"""

    def test_prefecture_order_北海道が0番(self):
        """北海道が0番であることを確認"""
        from processors.common.prefecture_order import get_prefecture_order
        assert get_prefecture_order("北海道") == 0

    def test_prefecture_order_沖縄県が46番(self):
        """沖縄県が46番であることを確認"""
        from processors.common.prefecture_order import get_prefecture_order
        assert get_prefecture_order("沖縄県") == 46

    def test_extract_prefecture_from_address(self):
        """住所から都道府県を正しく抽出できることを確認"""
        from processors.common.prefecture_order import extract_prefecture_from_address
        assert extract_prefecture_from_address("北海道札幌市北区") == "北海道"
        assert extract_prefecture_from_address("東京都新宿区") == "東京都"
        assert extract_prefecture_from_address("不明な住所") == ""


class TestFiltering:
    """フィルタリング処理のテスト"""

    def test_filter_records_5条件統合(self):
        """5つのフィルタ条件を満たすレコードのみが残ることを確認"""
        from processors.visit_list.processor import filter_records

        today = datetime.now().date()

        # 121列のダミーデータ作成（必要な列のみ実データ）
        data = {i: [None] * 6 for i in range(121)}
        data[0] = [1001, 1002, 1003, 1004, 1005, 1006]  # 管理番号
        data[86] = ["交渉困難", "死亡決定", "弁護士介入", "交渉継続中", "交渉困難", "追跡調査中"]  # 回収ランク
        data[72] = [today - timedelta(days=1), today, today + timedelta(days=1),
                    today - timedelta(days=5), today - timedelta(days=2), today + timedelta(days=3)]  # 入金予定日
        data[73] = [10, 2, 5, 10, 3, 15]  # 入金予定金額
        data[118] = ['5', None, '10', '5', '', '3']  # 委託先法人ID
        data[23] = ["北海道", "東京都", "", "大阪府", "神奈川県", "沖縄県"]  # 現住所1

        df = pd.DataFrame(data)
        filtered_df, logs = filter_records(df)

        # 期待: 1001のみ（交渉困難, 昨日, 10, '5', 北海道）
        assert len(filtered_df) == 1
        assert filtered_df.iloc[0, 0] == 1001


class TestDataExtraction:
    """データ抽出と22列マッピングのテスト"""

    def test_create_output_row_契約者(self):
        """ContractListの1行から22列の出力行を作成できることを確認"""
        from processors.visit_list.processor import create_output_row, VisitListConfig

        # 121列のダミーデータ作成
        data = {i: [None] for i in range(121)}
        data[0] = [1001]  # 管理番号
        data[2] = ["レントワン"]  # 最新契約種類
        data[14] = ["入居中"]  # 入居ステータス
        data[15] = ["滞納"]  # 滞納ステータス
        data[17] = [0]  # 退去手続き（実費）
        data[19] = ["営業太郎"]  # 営業担当者
        data[20] = ["田中太郎"]  # 契約者氏名
        data[23] = ["北海道"]  # 現住所1
        data[24] = ["札幌市"]  # 現住所2
        data[25] = ["北区1-1-1"]  # 現住所3
        data[71] = [50000]  # 滞納残債
        data[72] = ["2025-01-15"]  # 入金予定日
        data[73] = [10000]  # 入金予定金額
        data[84] = [80000]  # 月額賃料合計
        data[86] = ["交渉困難"]  # 回収ランク
        data[97] = ["C001"]  # クライアントCD
        data[98] = ["テスト管理会社"]  # クライアント名
        data[118] = ["5"]  # 委託先法人ID
        data[119] = ["テスト委託先"]  # 委託先法人名
        data[120] = ["2025-03-31"]  # 解約日

        df = pd.DataFrame(data)
        row = df.iloc[0]

        config = VisitListConfig.PERSON_TYPES["contractor"]
        output = create_output_row(row, "contractor", config)

        # 22列の検証
        assert output["管理番号"] == 1001
        assert output["最新契約種類"] == "レントワン"
        assert output["入居ステータス"] == "入居中"
        assert output["[人物]氏名"] == "田中太郎"  # シートによって変わるプレースホルダー
        assert output["現住所1"] == "北海道"
        assert output["回収ランク"] == "交渉困難"

    def test_create_output_row_結合住所(self):
        """結合住所列が正しく生成されることを確認"""
        from processors.visit_list.processor import create_output_row, VisitListConfig

        data = {i: [None] for i in range(121)}
        data[0] = [1001]
        data[23] = ["北海道"]
        data[24] = ["札幌市"]
        data[25] = ["北区1-1-1"]

        df = pd.DataFrame(data)
        row = df.iloc[0]

        config = VisitListConfig.PERSON_TYPES["contractor"]
        output = create_output_row(row, "contractor", config)

        # 9列目の空白ヘッダー（結合住所）を確認
        output_list = list(output.items())
        # 9列目（インデックス8）が結合住所であることを確認
        assert output_list[8][0] == ""  # 空白ヘッダー
        assert output_list[8][1] == "北海道札幌市北区1-1-1"  # 結合住所


class TestSorting:
    """ソート処理のテスト"""

    def test_sort_by_prefecture_北から南(self):
        """都道府県順（北→南）にソートされることを確認"""
        from processors.visit_list.processor import sort_by_prefecture, OUTPUT_COLUMNS

        # 5行のサンプルデータ（都道府県がバラバラ）
        df = pd.DataFrame([
            {col: None for col in OUTPUT_COLUMNS} for _ in range(5)
        ])
        df["管理番号"] = [1001, 1002, 1003, 1004, 1005]
        df["現住所1"] = ["沖縄県", "東京都", "北海道", "大阪府", "神奈川県"]

        sorted_df = sort_by_prefecture(df)
        result_ids = sorted_df["管理番号"].tolist()

        # 期待: 北海道(0) → 東京都(12) → 神奈川県(13) → 大阪府(26) → 沖縄県(46)
        assert result_ids == [1003, 1002, 1005, 1004, 1001]


class TestExcelGeneration:
    """Excel生成処理のテスト"""

    def test_generate_excel_5シート生成(self):
        """5シート（契約者/保証人1/保証人2/連絡人1/連絡人2）が生成されることを確認"""
        from processors.visit_list.processor import generate_excel, OUTPUT_COLUMNS
        import openpyxl

        # サンプルデータ（各シート1行ずつ）
        df_dict = {}
        for person_type in ["contractor", "guarantor1", "guarantor2", "contact1", "contact2"]:
            df = pd.DataFrame([{col: f"test_{person_type}" for col in OUTPUT_COLUMNS}])
            df_dict[person_type] = df

        excel_bytes, logs = generate_excel(df_dict, "test.xlsx")

        # Excelファイルを読み込んで検証
        import io as io_module
        wb = openpyxl.load_workbook(io_module.BytesIO(excel_bytes))

        # 5シート存在確認
        assert len(wb.sheetnames) == 5
        assert "契約者" in wb.sheetnames
        assert "保証人1" in wb.sheetnames

    def test_generate_excel_フォント設定(self):
        """游ゴシック 11ptが設定されることを確認"""
        from processors.visit_list.processor import generate_excel, OUTPUT_COLUMNS
        import openpyxl

        df_dict = {"contractor": pd.DataFrame([{col: "test" for col in OUTPUT_COLUMNS}])}
        excel_bytes, logs = generate_excel(df_dict, "test.xlsx")

        import io as io_module
        wb = openpyxl.load_workbook(io_module.BytesIO(excel_bytes))
        ws = wb["契約者"]

        # 最初のセルのフォントを確認
        cell = ws["A1"]
        assert cell.font.name == "游ゴシック"
        assert cell.font.size == 11
